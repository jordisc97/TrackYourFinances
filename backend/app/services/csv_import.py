import csv
import hashlib
import io
from datetime import date, datetime
import openpyxl
import xlrd
from sqlalchemy.orm import Session

from app.models import Account, Transaction, TransactionSource
from app.services.classification import classify_transaction, load_category_rules
from app.services.csv_mapping import SAMPLE_ROW_LIMIT, pick_mapped, resolve_column_mapping, _score_header, SCHEMA_FIELDS
from app.services.tx_enrichment import parse_card_merchant_location

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")
MISSING_MAPPING_MESSAGE = "Could not map CSV columns to booked_at and amount. Check headers or configure DEEPSEEK_API."


class CsvMappingError(ValueError):
    pass


def parse_date(value: str) -> date:
    cleaned = value.strip()[:10] if "T" in value else value.strip()
    for fmt in DATE_FORMATS:
        parsed = _safe_strptime(cleaned, fmt)
        if parsed is not None:
            return parsed
    raise ValueError(f"Unrecognized date: {value}")


def _safe_strptime(value: str, fmt: str) -> date | None:
    try:
        return datetime.strptime(value, fmt).date()
    except ValueError:
        return None


def parse_amount(value: str) -> float:
    cleaned = value.strip().replace("€", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".") if cleaned.rfind(",") > cleaned.rfind(".") else cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    return float(cleaned)

def _is_excel(filename: str, suffix: str) -> bool:
    return filename.lower().endswith(suffix)


MAX_HEADER_SCAN_ROWS = 20
MIN_HEADER_MATCHES = 2


def _normalize_header_value(value: object) -> str:
    return "" if value is None else str(value).strip()


def _row_looks_like_headers(cell_values: tuple | list) -> bool:
    non_empty = [_normalize_header_value(v) for v in cell_values if _normalize_header_value(v)]
    if len(non_empty) < MIN_HEADER_MATCHES:
        return False
    matched = sum(1 for header in non_empty if any(_score_header(field, header) > 0 for field in SCHEMA_FIELDS))
    return matched >= MIN_HEADER_MATCHES


def _parse_excel_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.active
    # Scan for the real header row (skip bank metadata/title rows).
    header_row_idx = 1
    for scan_row in sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True):
        if _row_looks_like_headers(scan_row):
            break
        header_row_idx += 1
    header_values = next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True), None) or ()
    headers: list[str] = []
    header_indexes: list[int] = []
    for col_idx, cell_value in enumerate(header_values):
        header = _normalize_header_value(cell_value)
        if header:
            headers.append(header)
            header_indexes.append(col_idx)
    rows: list[dict[str, str]] = []
    for excel_row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if excel_row is None:
            continue
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in excel_row):
            continue
        row_dict: dict[str, str] = {}
        for header, col_idx in zip(headers, header_indexes):
            value = excel_row[col_idx] if col_idx < len(excel_row) else None
            if value is None:
                row_dict[header] = ""
                continue
            if isinstance(value, datetime):
                row_dict[header] = value.date().isoformat()
            elif isinstance(value, date):
                row_dict[header] = value.isoformat()
            else:
                row_dict[header] = str(value).strip()
        rows.append(row_dict)
    return headers, rows


def _parse_excel_xls(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    # Scan for the real header row.
    header_row_idx = 0
    for scan_idx in range(min(sheet.nrows, MAX_HEADER_SCAN_ROWS)):
        row_vals = [sheet.cell_value(scan_idx, c) for c in range(sheet.ncols)]
        if _row_looks_like_headers(row_vals):
            header_row_idx = scan_idx
            break
    raw_headers: list[str] = []
    header_indexes: list[int] = []
    for col_idx in range(sheet.ncols):
        header = _normalize_header_value(sheet.cell_value(header_row_idx, col_idx))
        if header:
            raw_headers.append(header)
            header_indexes.append(col_idx)
    rows: list[dict[str, str]] = []
    for row_idx in range(header_row_idx + 1, sheet.nrows):
        row_dict: dict[str, str] = {}
        row_has_values = False
        for header, col_idx in zip(raw_headers, header_indexes):
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row_dict[header] = ""
                continue
            row_has_values = True
            if cell.ctype == xlrd.XL_CELL_DATE:
                dt_tuple = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
                row_dict[header] = date(dt_tuple[0], dt_tuple[1], dt_tuple[2]).isoformat()
            else:
                row_dict[header] = str(cell.value).strip()
        if not row_has_values:
            continue
        rows.append(row_dict)
    return raw_headers, rows


def _parse_csv_bytes(content_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    content = content_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    headers = list(reader.fieldnames or [])
    rows = [{key: (value or "") for key, value in row.items() if key is not None} for row in reader]
    return headers, rows


def _stable_external_id(account_id: int, row: dict, mapping, seen_keys: dict[str, int]) -> str:
    bank_id = pick_mapped(row, mapping, "external_id")
    if bank_id:
        base = bank_id
    else:
        fingerprint = "|".join([
            pick_mapped(row, mapping, "booked_at"),
            pick_mapped(row, mapping, "amount"),
            pick_mapped(row, mapping, "raw_description"),
            pick_mapped(row, mapping, "merchant"),
        ])
        digest = hashlib.sha256(f"{account_id}:{fingerprint}".encode()).hexdigest()[:32]
        base = f"csv-{digest}"
    duplicate_count = seen_keys.get(base, 0)
    seen_keys[base] = duplicate_count + 1
    return base if duplicate_count == 0 else f"{base}-dup{duplicate_count}"


def _parse_upload_to_headers_and_rows(filename: str, content_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        try:
            return _parse_excel_xlsx(content_bytes)
        except Exception:
            return _parse_excel_xls(content_bytes)
    if lower.endswith(".xls"):
        return _parse_excel_xls(content_bytes)
    return _parse_csv_bytes(content_bytes)


def import_transactions_upload(
    db: Session,
    account: Account,
    filename: str,
    content_bytes: bytes,
    overwrite: bool = False,
) -> tuple[int, int, int, int]:
    headers, rows = _parse_upload_to_headers_and_rows(filename, content_bytes)
    mapping = resolve_column_mapping(headers, rows[:SAMPLE_ROW_LIMIT])
    missing = mapping.missing_required()
    if missing:
        raise CsvMappingError(f"{MISSING_MAPPING_MESSAGE} Missing: {', '.join(missing)}")
    replaced = 0
    if overwrite:
        replaced = int(db.query(Transaction).filter(Transaction.account_id == account.id).delete(synchronize_session=False) or 0)
        db.commit()
    existing_ids: set[str] = set()
    if not overwrite:
        existing_ids = {row[0] for row in db.query(Transaction.external_id).filter(Transaction.account_id == account.id).all()}
    rules = load_category_rules(db, account.household_id)
    seen_keys: dict[str, int] = {}
    imported, skipped, categorized = 0, 0, 0
    for row in rows:
        date_raw = pick_mapped(row, mapping, "booked_at")
        amount_raw = pick_mapped(row, mapping, "amount")
        if not date_raw or not amount_raw:
            skipped += 1
            continue
        external_id = _stable_external_id(account.id, row, mapping, seen_keys)
        if not overwrite and external_id in existing_ids:
            skipped += 1
            continue
        description = pick_mapped(row, mapping, "raw_description")
        merchant = pick_mapped(row, mapping, "merchant")
        card_merchant, location = parse_card_merchant_location(description)
        tx = Transaction(
            account_id=account.id,
            booked_at=parse_date(date_raw),
            amount=parse_amount(amount_raw),
            currency=account.currency,
            raw_description=description,
            merchant=merchant or card_merchant or description,
            location=location,
            external_id=external_id,
            source=TransactionSource.csv.value,
        )
        classify_transaction(db, account.household_id, tx, rules=rules, use_llm=False)
        if tx.category_id is not None:
            categorized += 1
        db.add(tx)
        existing_ids.add(external_id)
        imported += 1
    db.commit()
    return imported, skipped, replaced, categorized
